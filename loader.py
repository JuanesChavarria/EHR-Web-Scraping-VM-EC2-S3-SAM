import datetime
import os
import pandas as pd
import openpyxl
from xls2xlsx import XLS2XLSX
import re
from datetime import datetime, timedelta, timezone
import pytz
from simple_salesforce import Salesforce
import io
import os
from dotenv import load_dotenv
import boto3

# Load environment variables from the .env file
load_dotenv()

# Retrieve Amazon credentials from environment variables
AWS_S3_BUCKET = os.getenv("AWS_S3_BUCKET")
AWS_ACCESS_KEY_ID = os.getenv("AWS_ACCESS_KEY_ID")
AWS_SECRET_ACCESS_KEY = os.getenv("AWS_SECRET_ACCESS_KEY")

s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY
)

# Retrieve Salesforce credentials from environment variables
sf_username = os.getenv('SF_USERNAME')
sf_password = os.getenv('SF_PASSWORD')
sf_security_token = os.getenv('SF_SECURITY_TOKEN')


def dir_xls2xlsx(dir_name):
    monthly_reports_xls = os.listdir(os.path.join(os.path.dirname(__file__), dir_name))
    for file in monthly_reports_xls:
        x2x = XLS2XLSX(os.path.join(os.path.dirname(__file__), dir_name, file))
        x2x.to_xlsx(os.path.join(os.path.dirname(__file__),dir_name,file + 'x'))
        os.remove(os.path.join(os.path.dirname(__file__),dir_name,file))

def get_patients_df():
    df_list = []
    patient_reports = os.listdir(os.path.join(os.path.dirname(__file__), 'PatientReports'))
    for file in patient_reports:
        workbook = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), 'PatientReports', file))
        sheet = workbook.active
        for merge in list(sheet.merged_cells):
            sheet.unmerge_cells(range_string = str(merge))
        sheet.delete_rows(1, amount=5)
        workbook.save(os.path.join("PatientReports",file))
        df_list.append(pd.read_excel(os.path.join(os.path.dirname(__file__),'PatientReports',file)))
    combined_df = pd.concat(df_list, axis=0)
    combined_df = combined_df.dropna(subset=['Last Name'])
    combined_df['PersonBirthdate'] = combined_df['Date of Birth'].apply(get_PersonBirthdate)
    combined_df['Phone'] = combined_df.apply(lambda row: get_Phone(row['Phone 1'], row['Phone 2'], row['Phone 3']), axis=1)
    combined_df['HealthCloudGA__Gender__pc'] = combined_df['Sex'].apply(get_HealthCloudGA__Gender__pc)
    combined_df['Source_System_Origin__c'] = 'AppointmentAndPatientNavigationSite'
    combined_df['EHR__c'] = 'a2r8b000000w4L9AAI'
    combined_df['RecordTypeId'] = '0128b000001BAccAAG'
    combined_df['HealthCloudGA__SourceSystemId__pc'] = combined_df['Chart Number']
    combined_df['FirstName'] = combined_df['First Name']
    combined_df['LastName'] = combined_df['Last Name']
    combined_df = combined_df.loc[:, ['HealthCloudGA__SourceSystemId__pc', 'EHR__c', 'HealthCloudGA__Gender__pc',
                'LastName', 'FirstName', 'Phone', 'PersonBirthdate', 'Source_System_Origin__c',
                'RecordTypeId']]
    combined_df.to_excel('PatientsToUpsert.xlsx', index=False)
    return combined_df
    
def patients_to_update_or_insert(patients):

    response_s3 = []

    sf = Salesforce(username=sf_username, password=sf_password, security_token=sf_security_token)
    results = sf.query_all("SELECT HealthCloudGA__SourceSystemId__pc, Id from Account WHERE AccountRecordTypeName__c = 'Patient' and HealthCloudGA__SourceSystem__pc = 'AppointmentAndPatientNavigationSite'")
    account_dict = {}
    for record in results['records']:
        account_dict[record['HealthCloudGA__SourceSystemId__pc']] = record['Id']
    patients['update_or_insert'] = patients['HealthCloudGA__SourceSystemId__pc'].apply(lambda x: check_if_update_or_insert(x, account_dict))
    update_df = patients.loc[patients['update_or_insert'] == 'update'].copy()  #update_df should be a datframe that contains all rows of patients where the 'update_or_insert' column has the value 'update'
    update_df['Id'] = update_df['HealthCloudGA__SourceSystemId__pc'].apply(lambda x: account_dict[x])
    update_df = update_df = update_df.loc[:, ['Id', 'HealthCloudGA__SourceSystemId__pc', 'EHR__c', 'HealthCloudGA__Gender__pc',
                'LastName', 'FirstName', 'Phone', 'PersonBirthdate', 'Source_System_Origin__c',
                'RecordTypeId']]
    update_df.to_excel("patients_updates.xlsx", index = False)

    with io.StringIO() as csv_buffer:
        update_df.to_csv(csv_buffer, index=False)

        response = s3_client.put_object(
            Bucket=AWS_S3_BUCKET, Key="patients/patients_updates.csv", Body=csv_buffer.getvalue()
        )

        print(response)
        response_s3.append(response)

    insert_df = patients.loc[patients['update_or_insert'] == 'insert']
    insert_df = insert_df.loc[:, ['HealthCloudGA__SourceSystemId__pc', 'EHR__c', 'HealthCloudGA__Gender__pc',
                'LastName', 'FirstName', 'Phone', 'PersonBirthdate', 'Source_System_Origin__c',
                'RecordTypeId']]
    insert_df.to_excel("patient_inserts.xlsx", index=False)

    with io.StringIO() as csv_buffer_2:
        insert_df.to_csv(csv_buffer_2, index=False)

        response = s3_client.put_object(
            Bucket=AWS_S3_BUCKET, Key="patients/patient_inserts.csv", Body=csv_buffer_2.getvalue()
        )

        print(response)
        response_s3.append(response)

    return response_s3

def check_if_update_or_insert(id, accounts_dict):
    if id in accounts_dict:
        return 'update'
    else:
        return 'insert'
    
def get_HealthCloudGA__Gender__pc(sex):
    if sex == 'F':
        return 'Female'
    elif sex == 'M':
        return 'Male'
    else:
        return ''

def get_PersonBirthdate(date):
    month, day, year = date.split('/')
    new_date = f"{year}-{month}-{day}"
    return new_date

def convert_phone_number(phone_number):
    new_phone_number = ''.join(filter(str.isdigit, phone_number))
    return new_phone_number

def get_Phone(phone_1, phone_2, phone_3):
    if type(phone_1) == str and len(phone_1) != 0:
        return convert_phone_number(phone_1)
    elif type(phone_2) == str and len(phone_2) != 0:
        return convert_phone_number(phone_2)
    elif type(phone_3) == str and len(phone_3) != 0:
        return convert_phone_number(phone_3)
    else:
        return ''
    
def get_monthly_appts_df():
    monthly_reports = os.listdir(os.path.join(os.path.dirname(__file__), 'MonthlyAppointmentReports'))
    df_list = []
    for file in monthly_reports:
        workbook = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), 'MonthlyAppointmentReports', file))
        sheet = workbook.active
        for merge in list(sheet.merged_cells):
            sheet.unmerge_cells(range_string=str(merge))
        sheet.insert_cols(1)
        office = None
        date = None
        for row in sheet.iter_rows(min_col=2, max_col=2, min_row=6):
            cell = row[0]
            date_pattern = r'^[A-Za-z]+,\s\d+/\d+/\d+$'
            if(cell.value != None):
                if cell.value.startswith("Office: "):
                    office = cell.value
                elif re.match(date_pattern, cell.value):
                    date = cell.value
            sheet.cell(row=cell.row, column=1).value = office
            sheet.cell(row=cell.row, column=2).value = date
        sheet.delete_rows(1,amount=4)
        workbook.save(os.path.join("MonthlyAppointmentReports",file))
        df_list.append(pd.read_excel(os.path.join(os.path.dirname(__file__),"MonthlyAppointmentReports",file)))
    combined_df = pd.concat(df_list, axis=0)
    combined_df = combined_df.rename(columns={'Unnamed: 0': 'Location'})
    combined_df = combined_df.loc[:, ~combined_df.columns.str.contains('^Unnamed')]
    combined_df = combined_df.dropna(subset=['Patient ID'])
    combined_df['ServiceTerritoryId'] = combined_df['Location'].apply(get_ServiceTerritoryId)
    combined_df['EHR__c'] = 'a2r8b000000w4L9AAI'
    combined_df['Duration'] = combined_df['Length'].apply(get_duration)
    combined_df['SchedStartTime'] = combined_df.apply(lambda row: get_SchedStartTime(row['Date'], row['Time']), axis=1)
    combined_df['Status'] = combined_df.apply(lambda row: get_status(row['Status'], row['Reason For Visit'], row['SchedStartTime']), axis=1)
    combined_df['Time'] = combined_df['Time'].apply(lambda x: x.strip())
    combined_df['Patient ID'] = combined_df['Patient ID'].apply(lambda x: str(int(x)))
    combined_df['ServiceResourceId'] = combined_df['Provider'].apply(get_ServiceResourceID)
    combined_df['isRequiredResource'] = 'TRUE'
    combined_df['WorkTypeId'] = combined_df.apply(lambda row: get_WorkTypeId(row['Reason For Visit'], row['Provider']), axis=1)
    print("about to get sched end times")

    sf = Salesforce(username=sf_username, password=sf_password, security_token=sf_security_token)
    results = sf.query_all("SELECT Id, EstimatedDuration FROM WorkType WHERE Active__c = TRUE")
    WorkType_dict = {}
    for record in results['records']:
        WorkType_dict[record['Id']] = int(record['EstimatedDuration'])
    combined_df['SchedEndTime'] = combined_df.apply(lambda row: get_SchedEndTime(row['Date'], row['Time'], row['WorkTypeId'], WorkType_dict), axis=1)
    print("got sched end times")
    combined_df['Description'] = ''
    combined_df['ContactId'] = combined_df['Provider'].apply(get_ContactId)
    PatientID_to_Parent_Record_Id = get_PatientId_to_ParentRecordId()
    combined_df['ParentRecordId'] = combined_df['Patient ID'].apply(lambda x: convert_by_dict(x, PatientID_to_Parent_Record_Id))
    return(combined_df)
    
def convert_by_dict(key, dict):
    if key in dict:
        return dict[key]
    else:
        return ''

def get_SchedStartTime(date, time):
    dt = datetime.strptime(f'{date}, {time}', '%A, %m/%d/%Y, %I:%M %p')
    pacific_tz = pytz.timezone('America/Los_Angeles')
    dt = pacific_tz.localize(dt)
    dt_utc = dt.astimezone(pytz.utc)
    dt_string = dt_utc.strftime('%Y-%m-%dT%H:%M:%S.%f')[:23] + "+0000"
    return dt_string

def get_SchedEndTime(date, time, WorkTypeId, WorkTypeId_to_duration) :
    try:
        length = WorkTypeId_to_duration[WorkTypeId]
    except KeyError:
        return ''

    dt = datetime.strptime(f'{date}, {time}', '%A, %m/%d/%Y, %I:%M %p')
    pacific_tz = pytz.timezone('America/Los_Angeles')
    dt = pacific_tz.localize(dt)
    dt = dt + timedelta(minutes = length)
    dt_utc = dt.astimezone(pytz.utc)
    dt_string = dt_utc.strftime('%Y-%m-%dT%H:%M:%S.%f')[:23] + "+0000"
    return dt_string

def get_status(status, reason_for_visit, SchedStartTime):
    OA_status_to_SF = {'Active' : 'Scheduled',
                'Cancelled by Office' : 'Canceled',
                'Cancelled by Patient' : 'Canceled',
                'Cancelled by Provider' : 'Canceled',
                'Checked In' : 'Scheduled',
                'Checked Out' : 'Scheduled',
                'Completed' : 'Scheduled',
                'Confirmed' : 'Scheduled',
                'In Room' : 'Scheduled',
                'Last Minute Cancel by Patient' : 'Canceled',
                'Last Minute Reschedule' : 'Canceled',
                'Left Message' : 'Scheduled',
                'Patient Did Not Come' : 'Canceled',
                'Rescheduled' : 'Canceled',
                'Visit Created' : 'Scheduled'}
    if(OA_status_to_SF[status] == 'Canceled' or ('cancel' in reason_for_visit.lower())):
        return 'Canceled'
    else:
            if datetime.strptime(SchedStartTime, '%Y-%m-%dT%H:%M:%S.%f%z') < datetime.now(timezone.utc):
                return 'Completed'
            else:
                return 'Scheduled'

def get_duration(length):
    return length.split(" ")[0]

def get_ContactId(provider):
    provider_to_ContactId =  {'InactiveDoctor 1, MD' : '',
                            'ActiveDoctor 1, MD' : '0038b00002zgd4MAAQ',
                            'ActiveDoctor 2, MD' : '0038b00002zgd4LAAQ',
                            'ActiveDoctor 3, MD' : '0038b00002zgd4PAAQ',
                            'InactiveDoctor 2, MD' : '',
                            'InactiveDoctor 3, MD'  : '',
                            'InactiveDoctor 4' : '',
                            'InactiveDoctor 5, MD' : ''}
    return provider_to_ContactId[provider]


def get_ServiceResourceID(provider):
    provider_to_service_resource_id = {'InactiveDoctor 1, MD' : '',
                                       'ActiveDoctor 1, MD' : '0Hn8b0000008yQUCAY',
                                       'ActiveDoctor 2, MD' : '0Hn8b0000008yQoCAI',
                                       'ActiveDoctor 3, MD' : '0Hn8b0000008yQkCAI',
                                       'InactiveDoctor 2, MD' : '',
                                       'InactiveDoctor 3, MD'  : '',
                                       'InactiveDoctor 4' : '',
                                       'InactiveDoctor 5, MD' : ''}
    return provider_to_service_resource_id[provider]

def get_WorkTypeId(reason_for_vist, provider):

    WorkTypeName_to_Id = {
        'Patch Test': '08q8b0000008uu6AAA',
        '-- Physical Exam --': '08q8b0000008uu7AAA',
        'PPD Reading': '08q8b0000008uu9AAA',
        'Pre-Op': '08q8b0000008uuAAAQ',
        'Shot(s)': '08q8b0000008uuCAAQ',
        'Pap Smear': '08q8b0000008uu5AAA',
        '-- Sick Visit --': '08q8b0000008uuDAAQ',
        'Skin Test': '08q8b0000008uuEAAQ',
        '-- New Patient Consultation --': '08q8b0000008uuFAAQ',
        '-- New Patient (PCP) --': '08q8b0000008uu0AAA',
        'Initial Allergy Shot': '08q8b0000008utzAAA',
        '-- Existing Patient Follow Up (PCP & Endocrinology) --': '08q8b0000008utmAAA',
        'Walk In': '08q8b0000008utoAAA',
        'Challenge Test': '08q8b0000008utpAAA',
        'Telehealth Consult': '08q8b0000008utsAAA',
        'Telehealth Follow Up': '08q8b0000008uttAAA',
        'FU Allergy Shot': '08q8b0000008utvAAA',
        '-- Existing Patient Follow Up (Dr. 2 and Dr. 3) --': '08q8b0000008utlAAA',
        '-- Medicare Annual Wellness Visit --': '08q8b0000008uzuAAA',
        '-- New Patient Consultation (Dr. 2 and Dr. 3) --': '08q8b0000008vJ1AAI',
        '-- Existing Patient Follow Up (Dr. 1) --': '08q8b0000008vJ6AAI',
        '-- New Patient Consultation (Dr. 1) --': '08q8b0000008vIwAAI'
    }

    if provider == 'ActiveDoctor 2, MD' or provider == 'ActiveDoctor 3, MD':
        if ('follow up' in reason_for_vist.lower()) or ('f/u' in reason_for_vist.lower()):
            return WorkTypeName_to_Id['-- Existing Patient Follow Up (Dr. 2 and Dr. 3) --']
        else:
            return WorkTypeName_to_Id['-- New Patient Consultation (Dr. 2 and Dr. 3) --']
    elif provider == 'ActiveDoctor 1, MD':
         if ('follow up' in reason_for_vist.lower()) or ('f/u' in reason_for_vist.lower()):
             return WorkTypeName_to_Id['-- Existing Patient Follow Up (Dr. 1) --']
         else:
             return WorkTypeName_to_Id['-- New Patient Consultation (Dr. 1) --']
    else:
        return ''

    
def get_ServiceTerritoryId(location):
    OA_location_to_SF_ServiceTerritoryName = {'Office: Clinic Location Name 1 office' : 'Clinic Location 1',
                                              'Clinic Location Name 2 office' : 'Clinic Location 2',
                                              'Office: Clinic Location Name 3' : 'Clinic Location 3',
                                              'Clinic Location Name 4'	: 'Clinic Location 4',
                                              'Office: Clinic Location Name 5' :	'Clinic Location 5',
                                              'Clinic Location Name 6' : 'Clinic Location 6',
                                              'Office: Clinic Location Name 7' : 'Clinic Location 7',
                                              'Office: Clinic Location Name 8' : 'Clinic Location 8',
                                              'Clinic Location Name 9' : 'Clinic Location 9',
                                              'Clinic Location Name 10' : 'Clinic Location 10',
                                              'Clinic Location Name 11' : 'Clinic Location 11',
                                              'Office: Clinic Location Name 12' : 'Clinic Location 12'}
    
    SF_ServiceTerritoryName_to_SF_ServiceTerritoryId = {'Clinic Location 1': '0Hh8b000000Y3PoCAK',
                                                        'Clinic Location 2': '0Hh8b000000Y3PpCAK',
                                                        'Clinic Location 3': '0Hh8b000000Y3PqCAK',
                                                        'Clinic Location 4': '0Hh8b000000Y3PrCAK',
                                                        'Clinic Location 5': '0Hh8b000000Y3PsCAK',
                                                        'Clinic Location 6': '0Hh8b000000Y3PtCAK',
                                                        'Clinic Location 7': '0Hh8b000000Y3PuCAK',
                                                        'Clinic Location 8': '0Hh8b000000Y3PvCAK',
                                                        'Clinic Location 9': '0Hh8b000000Y3PwCAK',
                                                        'Clinic Location 10': '0Hh8b000000Y3PxCAK',
                                                        'Clinic Location 11': '0Hh8b000000Y3PyCAK',
                                                        'Clinic Location 12': '0Hh8b000000Y3PzCAK'}
    return SF_ServiceTerritoryName_to_SF_ServiceTerritoryId[OA_location_to_SF_ServiceTerritoryName[location]]


def get_PatientId_to_ParentRecordId():
    sf = Salesforce(username=sf_username, password=sf_password, security_token=sf_security_token)
    results = sf.query_all("SELECT HealthCloudGA__SourceSystemId__pc, Id from Account WHERE AccountRecordTypeName__c = 'Patient' and HealthCloudGA__SourceSystem__pc = 'AppointmentAndPatientNavigationSite'")
    PatientId_to_ParentRecordId = {}
    for record in results['records']:
        PatientId_to_ParentRecordId[record['HealthCloudGA__SourceSystemId__pc']] = record['Id']
    return PatientId_to_ParentRecordId


def float_to_str(x):
    try:
        return str(int(x))
    except:
        return ''
def get_AssignedResource():
    assigned_resource_df = pd.read_excel(os.path.join(os.path.dirname(__file__),'pass_to_service_resource.xlsx'))
    assigned_resource_df['EHRAppointmentId__c'] = assigned_resource_df['EHRAppointmentId__c'].apply(float_to_str)
    sf = Salesforce(username=sf_username, password=sf_password, security_token=sf_security_token)
    results = sf.query_all("SELECT EHRAppointmentId__c, Id FROM ServiceAppointment WHERE EHR_Name__c  = 'AppointmentAndPatientNavigationSite'")
    records = results['records']
    result_dict = {}
    for record in records:
        ehr_appointment_id = record['EHRAppointmentId__c']
        appointment_id = record['Id']
        result_dict[ehr_appointment_id] = appointment_id
    assigned_resource_df['ServiceAppointmentId'] = assigned_resource_df['EHRAppointmentId__c'].apply(lambda x: convert_by_dict(x,result_dict))
    assigned_resource_df['IsRequiredResource'] = 'TRUE'
    assigned_resource_df.to_excel('AssignedResource_upserts.xlsx', index=False)

def main():
    patients_to_update_or_insert(None)

if __name__ == "__main__":
    main()